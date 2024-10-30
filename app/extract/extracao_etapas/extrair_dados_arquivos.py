import requests
import io
import pandas as pd
from openpyxl import load_workbook
from odf.opendocument import load
from odf.table import Table
from odf.table import TableRow, TableCell
from odf.text import P

def acessar_arquivos_extrair_dados(link_base: str, formato_arquivo: str) -> pd.DataFrame:
    try:
        # Lista de nomes de abas a serem buscadas
        abas_alvo = ['por_Orgao_e_Cargo', 'por Orgao e Cargo', 'por_Órgao_e_Cargo', 'por Órgao e Cargo', 'por_Órgão_e_Cargo']

        if formato_arquivo == "ods":
            response_ods = requests.get(f"{link_base}.ods")
            ods_file = io.BytesIO(response_ods.content)
            document = load(ods_file)
            sheets = [table.getAttribute("name") for table in document.getElementsByType(Table)]

            for sheet_name in sheets:
                if sheet_name in abas_alvo:
                    print(f"Carregando a aba '{sheet_name}' do arquivo .ods")
                    table = next(table for table in document.getElementsByType(Table) if table.getAttribute("name") == sheet_name)
                    rows = []
                    for row in table.getElementsByType(TableRow):
                        cells = []
                        for cell in row.getElementsByType(TableCell):
                            # Coleta o texto de cada célula (dentro dos elementos <text:p>)
                            cell_text = "".join([node.data for p in cell.getElementsByType(P) for node in p.childNodes if node.nodeType == node.TEXT_NODE])
                            cells.append(cell_text)
                        rows.append(cells)
                    df = pd.DataFrame(rows[1:], columns=rows[0])
                    

        elif formato_arquivo == "xlsx":
            response_xlsx = requests.get(f"{link_base}.xlsx")
            xlsx_file = io.BytesIO(response_xlsx.content)
            workbook = load_workbook(xlsx_file, read_only=True)

            for sheet_name in workbook.sheetnames:
                if sheet_name in abas_alvo:
                    print(f"Carregando a aba '{sheet_name}' do arquivo .xlsx")
                    df = pd.read_excel(xlsx_file, sheet_name=sheet_name)
                

    except requests.HTTPError as e:
        print(f"Erro ao acessar o link: {e}")
        return "erro"
    except Exception as e:
        print(f"Erro ao processar o arquivo: {e}")
        return "erro"
    
    return df
import requests
import io
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from odf.opendocument import load
from odf.table import Table
from odf.table import TableRow, TableCell
from .contract import validar_dados
from odf.text import P

def verificar_tipo_arquivo(data):
    # Define os três links base possíveis
    """
    Verifica se o link do arquivo existe e qual é o tipo (ods ou xlsx) e
    retorna o link base, tipo de link e formato de arquivo.

    Se o arquivo não for encontrado, retorna "arquivo não encontrado".
    Se houver um erro ao acessar o link, retorna "erro".
    Se houver um erro ao processar o arquivo, retorna "erro".

    Parameters
    ----------
    data : str
        Data no formato "YYYYMM" da qual se deseja verificar a existência do arquivo.

    Returns
    -------
    str
        Link base, tipo de link e formato de arquivo, ou "arquivo não encontrado" ou "erro".
    """
    link_opcao1 = f"https://repositorio.dados.gov.br/segrt/LotOrgao_DistOcupVagas%20-%20{data}"
    link_opcao2 = f"https://repositorio.dados.gov.br/segrt/cargos%20vagos%20e%20vacancia/CargosVagosVacancias_{data}"
    link_opcao3 = f"https://repositorio.dados.gov.br/segrt/LotOrgao_DistOcupVagas_{data}"
    
    # Lista de opções para verificar
    opcoes = [
        (link_opcao1, "Opção 1"),
        (link_opcao2, "Opção 2"),
        (link_opcao3, "Opção 3")
    ]
    
    link_base = None
    formato_arquivo = None

    # Percorre as opções e verifica se o arquivo existe (ods ou xlsx)
    for link, tipo in opcoes:
        response_ods = requests.get(f"{link}.ods")
        response_xlsx = requests.get(f"{link}.xlsx")
        
        if response_ods.status_code == 200:
            link_base = link
            formato_arquivo = "ods"
            break
        elif response_xlsx.status_code == 200:
            link_base = link
            formato_arquivo = "xlsx"
            break
    
    # Caso nenhum link tenha sido encontrado
    if not link_base:
        print(f"Nenhum arquivo encontrado para {data}.")
        return "arquivo não encontrado"

    try:
        # Lista de nomes de abas a serem buscadas
        abas_alvo = ['por_Orgao_e_Cargo', 'por Orgao e Cargo', 'por_Órgao_e_Cargo', 'por Órgao e Cargo']

        if formato_arquivo == "ods":
            response_ods = requests.get(f"{link_base}.ods")
            ods_file = io.BytesIO(response_ods.content)
            document = load(ods_file)
            sheets = [table.getAttribute("name") for table in document.getElementsByType(Table)]

            for sheet_name in sheets:
                if sheet_name in abas_alvo:
                    print(f"Carregando a aba '{sheet_name}' do arquivo .ods")
                    table = next(table for table in document.getElementsByType(Table) if table.getAttribute("name") == sheet_name)
                    rows = []
                    for row in table.getElementsByType(TableRow):
                        cells = []
                        for cell in row.getElementsByType(TableCell):
                            # Coleta o texto de cada célula (dentro dos elementos <text:p>)
                            cell_text = "".join([node.data for p in cell.getElementsByType(P) for node in p.childNodes if node.nodeType == node.TEXT_NODE])
                            cells.append(cell_text)
                        rows.append(cells)
                    df = pd.DataFrame(rows[1:], columns=rows[0])
                    adicionar_metadados_e_salvar(df, data, formato_arquivo = 'ods')

        elif formato_arquivo == "xlsx":
            response_xlsx = requests.get(f"{link_base}.xlsx")
            xlsx_file = io.BytesIO(response_xlsx.content)
            workbook = load_workbook(xlsx_file, read_only=True)

            for sheet_name in workbook.sheetnames:
                if sheet_name in abas_alvo:
                    print(f"Carregando a aba '{sheet_name}' do arquivo .xlsx")
                    df = pd.read_excel(xlsx_file, sheet_name=sheet_name)
                    adicionar_metadados_e_salvar(df, data, formato_arquivo = 'xlsx')

    except requests.HTTPError as e:
        print(f"Erro ao acessar o link: {e}")
        return "erro"
    except Exception as e:
        print(f"Erro ao processar o arquivo: {e}")
        return "erro"


def adicionar_metadados_e_salvar(df, data, formato_arquivo):
    # Adiciona os metadados ao DataFrame
    """
    Adiciona metadados ao DataFrame fornecido e salva-o como um arquivo CSV.

    Esta função insere várias colunas de metadados no DataFrame, como 'Cliente', 'Tema',
    'Fonte', 'Data Carga', entre outros. Em seguida, valida os dados do DataFrame utilizando
    a função `validar_dados` e salva o DataFrame como um arquivo CSV nomeado com a data fornecida.

    Parâmetros:
    - df (pd.DataFrame): O DataFrame ao qual os metadados serão adicionados.
    - data (str): String representando a data, utilizada para nomear o arquivo CSV de saída.

    Retorno:
    - None: A função salva o DataFrame como um arquivo CSV no diretório atual e não retorna nada.
    """
    df['Cliente'] = "governo federal"
    df['Tema'] = "cargos e vacancias"
    df['Fonte'] = "https://dados.gov.br/dados/conjuntos-dados/gestao-de-pessoas-executivo-federal---cargos-vagos-e-vacancias"
    df['Data Carga'] = datetime.now().strftime('%Y-%m-%d')
    df['Origem Dados'] = formato_arquivo
    df['Periodicidade'] = "Mensal"
    df['Responsável'] = "Automação de Script"
    df['Versão Script'] = "1.0"
    df['Ambiente Execução'] = "Produção"

    # Validar o DataFrame usando pydantic
    validar_dados(df)

    # Salva o DataFrame como CSV
    try: 
        #Esse caminho está fixado, porém precisa ser dinamico conforme o ambiente em que estiver executando o script          
        nome_arquivo = f"C:/Users/smart/Documents/1.Estudos-Profissional/projeto-gestao-pessoas-exec-federal/data/{data}.csv"
        df.to_csv(nome_arquivo, index=False, encoding='utf-8-sig', sep=';')
        print(f"Arquivo salvo como {nome_arquivo}")
    except Exception as e:
        print(f"Erro ao salvar o arquivo: {e}")

################################### Execução das funções ############################################
if __name__ == "__main__":
    # Exemplo de uso
    lista_datas = [
        '201601', '201602', '201603', '201604', '201605', '201606', '201607', '201608', '201609', '201610', '201611', '201612',
        '201701', '201702', '201703', '201704', '201705', '201706', '201707', '201708', '201709', '201710', '201711', '201712',
        '201801', '201802', '201803', '201804', '201805', '201806', '201807', '201808', '201809', '201810', '201811', '201812',
        '201901', '201902', '201903', '201904', '201905', '201906', '201907', '201908', '201909', '201910', '201911', '201912',
        '202001', '202002', '202003', '202004', '202005', '202006', '202007', '202008', '202009', '202010', '202011', '202012',
        '202101', '202102', '202103', '202104', '202105', '202106', '202107', '202108', '202109', '202110', '202111', '202112',
        '202201', '202202', '202203', '202204', '202205', '202206', '202207', '202208', '202209', '202210', '202211', '202212',
        '202301', '202302', '202303', '202304', '202305', '202306', '202307', '202308', '202309', '202310', '202311', '202312',
        '202401', '202402', '202403', '202404', '202405', '202406', '202407', '202408', '202409', '202410', '202411', '202412',
    ]
    for data in lista_datas:
        tipo_arquivo = verificar_tipo_arquivo(data)